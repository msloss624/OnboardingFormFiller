"""
Excel generator — fills the RFI template with extracted answers,
color-codes cells by confidence, and adds a sources metadata sheet.
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from extraction.extractor import ExtractedAnswer
from schema.rfi_fields import RFI_FIELDS, Confidence


# Confidence → cell fill color
CONFIDENCE_FILLS = {
    Confidence.HIGH: PatternFill("solid", fgColor="C6EFCE"),      # Green
    Confidence.MEDIUM: PatternFill("solid", fgColor="FFEB9C"),    # Yellow
    Confidence.LOW: PatternFill("solid", fgColor="FFC7CE"),       # Pink/Red
    Confidence.MISSING: PatternFill("solid", fgColor="D9D9D9"),   # Gray
}

CONFIDENCE_FONTS = {
    Confidence.HIGH: Font(color="006100"),
    Confidence.MEDIUM: Font(color="9C5700"),
    Confidence.LOW: Font(color="9C0006"),
    Confidence.MISSING: Font(color="808080", italic=True),
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_rfi_excel(
    answers: list[ExtractedAnswer],
    template_path: str | Path,
    output_path: str | Path,
    company_name: str = "",
) -> dict:
    """
    Fill the RFI template with extracted answers.

    Returns stats about completion.
    """
    template_path = Path(template_path)
    output_path = Path(output_path)

    wb = load_workbook(template_path)
    ws = wb.active  # "RFI " sheet

    # Response column = B, Notes column = C
    response_col = 2  # B
    notes_col = 3     # C

    filled = 0
    total = len(answers)
    by_confidence = {c: 0 for c in Confidence}

    for answer in answers:
        row = answer.row
        confidence = answer.confidence
        by_confidence[confidence] += 1

        # Fill response cell
        response_cell = ws.cell(row=row, column=response_col)
        if answer.answer:
            response_cell.value = answer.answer
            filled += 1
        else:
            response_cell.value = "— Not found —"

        # Apply confidence coloring
        response_cell.fill = CONFIDENCE_FILLS[confidence]
        response_cell.font = CONFIDENCE_FONTS[confidence]
        response_cell.alignment = Alignment(wrap_text=True, vertical="top")
        response_cell.border = THIN_BORDER

        # Fill notes cell with source + evidence
        notes_cell = ws.cell(row=row, column=notes_col)
        notes_parts = []
        if answer.source:
            notes_parts.append(f"Source: {answer.source}")
        if answer.evidence:
            notes_parts.append(f'Evidence: "{answer.evidence}"')
        notes_cell.value = "\n".join(notes_parts) if notes_parts else ""
        notes_cell.alignment = Alignment(wrap_text=True, vertical="top")
        notes_cell.border = THIN_BORDER

    # Set column widths
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50

    # Add metadata/legend sheet
    _add_metadata_sheet(wb, answers, company_name)

    wb.save(output_path)

    return {
        "total_fields": total,
        "filled": filled,
        "completion_pct": round(filled / total * 100, 1) if total > 0 else 0,
        "by_confidence": {c.value: count for c, count in by_confidence.items()},
        "output_path": str(output_path),
    }


def _add_metadata_sheet(wb, answers: list[ExtractedAnswer], company_name: str):
    """Add a metadata sheet with legend, sources, and completion stats."""
    if "Metadata" in wb.sheetnames:
        del wb["Metadata"]
    ws = wb.create_sheet("Metadata")

    # Header
    ws["A1"] = "Onboarding Form Filler — Generation Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Company: {company_name}"
    ws["A2"].font = Font(size=12)

    # Legend
    ws["A4"] = "Confidence Legend"
    ws["A4"].font = Font(bold=True, size=11)
    legend = [
        (Confidence.HIGH, "High — explicitly stated with specific details"),
        (Confidence.MEDIUM, "Medium — mentioned but vague, or inferred from context"),
        (Confidence.LOW, "Low — indirect reference, educated guess"),
        (Confidence.MISSING, "Missing — not found in any source"),
    ]
    for i, (conf, desc) in enumerate(legend):
        row = 5 + i
        cell = ws.cell(row=row, column=1, value=conf.value.title())
        cell.fill = CONFIDENCE_FILLS[conf]
        cell.font = Font(bold=True)
        ws.cell(row=row, column=2, value=desc)

    # Stats
    ws["A10"] = "Completion Stats"
    ws["A10"].font = Font(bold=True, size=11)
    total = len(answers)
    filled = sum(1 for a in answers if a.answer)
    ws["A11"] = f"Total fields: {total}"
    ws["A12"] = f"Filled: {filled}"
    ws["A13"] = f"Completion: {round(filled/total*100, 1) if total else 0}%"

    # Sources used
    ws["A15"] = "Sources"
    ws["A15"].font = Font(bold=True, size=11)
    sources = sorted(set(a.source for a in answers if a.source))
    for i, src in enumerate(sources):
        ws.cell(row=16 + i, column=1, value=src)

    # Detailed per-field breakdown
    detail_start = 16 + len(sources) + 2
    ws.cell(row=detail_start, column=1, value="Field").font = Font(bold=True)
    ws.cell(row=detail_start, column=2, value="Confidence").font = Font(bold=True)
    ws.cell(row=detail_start, column=3, value="Source").font = Font(bold=True)
    ws.cell(row=detail_start, column=4, value="Evidence").font = Font(bold=True)

    for i, a in enumerate(answers):
        r = detail_start + 1 + i
        ws.cell(row=r, column=1, value=a.question)
        conf_cell = ws.cell(row=r, column=2, value=a.confidence.value)
        conf_cell.fill = CONFIDENCE_FILLS[a.confidence]
        ws.cell(row=r, column=3, value=a.source)
        ws.cell(row=r, column=4, value=a.evidence)

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 60


def load_answers_from_excel(path: str | Path) -> list[ExtractedAnswer]:
    """
    Load previously generated answers from an Excel file.

    Reads the RFI sheet column B for answers and the Metadata sheet
    for per-field confidence, source, and evidence.
    """
    wb = load_workbook(path, data_only=True)
    ws_rfi = wb.active
    ws_meta = wb["Metadata"] if "Metadata" in wb.sheetnames else None

    # Build per-field metadata from the Metadata sheet detail breakdown
    meta_lookup: dict[str, dict] = {}  # question -> {confidence, source, evidence}
    if ws_meta:
        # Find the detail header row ("Field" in column A)
        for row in ws_meta.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value == "Field":
                detail_start = cell.row
                # Read rows after the header
                for r in range(detail_start + 1, ws_meta.max_row + 1):
                    question = ws_meta.cell(row=r, column=1).value
                    conf_val = ws_meta.cell(row=r, column=2).value
                    source = ws_meta.cell(row=r, column=3).value or ""
                    evidence = ws_meta.cell(row=r, column=4).value or ""
                    if question:
                        meta_lookup[question] = {
                            "confidence": conf_val or "missing",
                            "source": source,
                            "evidence": evidence,
                        }
                break

    answers = []
    for field in RFI_FIELDS:
        cell_value = ws_rfi.cell(row=field.row, column=2).value  # Column B
        answer_text = str(cell_value).strip() if cell_value else None

        # Skip "— Not found —" placeholders
        if answer_text and answer_text == "— Not found —":
            answer_text = None

        meta = meta_lookup.get(field.question, {})
        conf_str = meta.get("confidence", "missing")
        try:
            confidence = Confidence(conf_str)
        except ValueError:
            confidence = Confidence.MISSING if not answer_text else Confidence.MEDIUM

        # If we have an answer but confidence says missing, correct it
        if answer_text and confidence == Confidence.MISSING:
            confidence = Confidence.MEDIUM

        # If no answer, force missing
        if not answer_text:
            confidence = Confidence.MISSING

        answers.append(ExtractedAnswer(
            field_key=field.key,
            question=field.question,
            answer=answer_text,
            confidence=confidence,
            source=meta.get("source", "Previous run"),
            evidence=meta.get("evidence", ""),
            row=field.row,
        ))

    wb.close()
    return answers


def get_previous_sources(path: str | Path) -> set[str]:
    """
    Read the Metadata sheet 'Sources' section and return the set of
    source names used in a previous generation.
    """
    wb = load_workbook(path, data_only=True)
    sources: set[str] = set()

    if "Metadata" not in wb.sheetnames:
        wb.close()
        return sources

    ws = wb["Metadata"]
    # Find the "Sources" header (in column A)
    in_sources = False
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value == "Sources":
            in_sources = True
            continue
        if in_sources:
            # Stop at next bold header or empty cell
            if not cell.value or (cell.font and cell.font.bold and cell.value != "Sources"):
                break
            sources.add(str(cell.value).strip())

    wb.close()
    return sources
